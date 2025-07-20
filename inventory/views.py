import csv
import openpyxl
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from reportlab.pdfgen import canvas
from .models import Ingredient, InventoryItem, InventoryTransaction
from django.utils.timezone import localtime

def ingredient(request):
    queryset = Ingredient.objects.all()
    search = request.GET.get("search")

    if search:
        queryset = queryset.filter(name__icontains=search)

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="ingredients.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Unit'])
        for obj in queryset:
            writer.writerow([obj.name, obj.unit])
        return response

    if request.GET.get("export") == "excel":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Unit'])
        for obj in queryset:
            sheet.append([obj.name, obj.unit])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="ingredients.xlsx"'
        workbook.save(response)
        return response

    if request.GET.get("export") == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="ingredients.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.drawString(100, y, 'Ingredients')
        y -= 20
        for obj in queryset:
            p.drawString(100, y, f"{obj.name} ({obj.unit})")
            y -= 20
        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    if request.method == "POST":
        obj_id = request.POST.get("id")
        name = request.POST.get("name")
        unit = request.POST.get("unit")

        if obj_id:
            ingredient = get_object_or_404(Ingredient, pk=obj_id)
            ingredient.name = name
            ingredient.unit = unit
            ingredient.save()
        else:
            Ingredient.objects.create(name=name, unit=unit)

        return redirect("ingredients")

    context = {
        "page_title": "Ingredients",
        "table_headers": ["Name", "Unit"],
        "table_rows": [[i.name, i.unit, i.id] for i in queryset],
    }
    return render(request, "inventory/ingredient.html", context)

def delete_ingredient(request, pk):
    get_object_or_404(Ingredient, pk=pk).delete()
    return redirect("ingredients")

def inventory(request):
    queryset = InventoryItem.objects.select_related("ingredient")
    search = request.GET.get("search")

    if search:
        queryset = queryset.filter(ingredient__name__icontains=search)

    export = request.GET.get("export")
    if export == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory.csv"'
        writer = csv.writer(response)
        writer.writerow(['Ingredient', 'Stock', 'Reorder Level'])
        for obj in queryset:
            writer.writerow([obj.ingredient.name, obj.quantity_in_stock, obj.reorder_level])
        return response

    elif export == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Ingredient', 'Stock', 'Reorder Level'])
        for obj in queryset:
            ws.append([obj.ingredient.name, float(obj.quantity_in_stock), float(obj.reorder_level)])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inventory.xlsx"'
        wb.save(response)
        return response

    elif export == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="inventory.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.drawString(100, y, "Inventory Report")
        y -= 20
        for obj in queryset:
            p.drawString(100, y, f"{obj.ingredient.name} - {obj.quantity_in_stock} (Reorder: {obj.reorder_level})")
            y -= 15
        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    if request.method == "POST":
        obj_id = request.POST.get("id")
        ingredient_id = request.POST.get("ingredient")
        quantity = request.POST.get("quantity")
        reorder = request.POST.get("reorder")

        if obj_id:
            item = get_object_or_404(InventoryItem, pk=obj_id)
            item.ingredient_id = ingredient_id
            item.quantity_in_stock = quantity
            item.reorder_level = reorder
            item.save()
        else:
            InventoryItem.objects.create(
                ingredient_id=ingredient_id,
                quantity_in_stock=quantity,
                reorder_level=reorder
            )
        return redirect("inventory_items")

    context = {
        "page_title": "Inventory",
        "table_headers": ["Ingredient", "Stock Qty", "Reorder Level"],
        "table_rows": [
            [i.ingredient.name, i.quantity_in_stock, i.reorder_level, i.id] for i in queryset
        ],
        "ingredients": Ingredient.objects.all()
    }
    return render(request, "inventory/inventory.html", context)

def delete_inventory(request, pk):
    get_object_or_404(InventoryItem, pk=pk).delete()
    return redirect("inventory_items")

def inventory_transactions(request):
    transactions = InventoryTransaction.objects.select_related('inventory_item', 'inventory_item__ingredient')
    search = request.GET.get("search")

    if search:
        transactions = transactions.filter(inventory_item__ingredient__name__icontains=search)

    export = request.GET.get("export")
    if export == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
        writer = csv.writer(response)
        writer.writerow(['Ingredient', 'Change', 'Type', 'Date', 'Notes'])
        for t in transactions:
            writer.writerow([t.inventory_item.ingredient.name, t.quantity_change, t.get_transaction_type_display(), localtime(t.date).strftime('%Y-%m-%d %H:%M'), t.notes])
        return response

    elif export == "excel":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Ingredient', 'Change', 'Type', 'Date', 'Notes'])
        for t in transactions:
            sheet.append([t.inventory_item.ingredient.name, t.quantity_change, t.get_transaction_type_display(), localtime(t.date).strftime('%Y-%m-%d %H:%M'), t.notes])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
        workbook.save(response)
        return response

    elif export == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="transactions.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.drawString(100, y, 'Inventory Transactions')
        y -= 20
        for t in transactions:
            p.drawString(100, y, f"{t.inventory_item.ingredient.name} - {t.quantity_change} ({t.get_transaction_type_display()})")
            y -= 15
        p.showPage()
        p.save()
        response.write(buffer.getvalue())
        buffer.close()
        return response

    if request.method == "POST":
        id = request.POST.get("id")
        item_id = request.POST.get("inventory_item")
        quantity = request.POST.get("quantity_change")
        ttype = request.POST.get("transaction_type")
        notes = request.POST.get("notes")

        if id:
            t = get_object_or_404(InventoryTransaction, pk=id)
            t.inventory_item_id = item_id
            t.quantity_change = quantity
            t.transaction_type = ttype
            t.notes = notes
            t.save()
        else:
            InventoryTransaction.objects.create(
                inventory_item_id=item_id,
                quantity_change=quantity,
                transaction_type=ttype,
                notes=notes,
            )
        return redirect("inventory_trans")

    context = {
        "page_title": "Inventory Transactions",
        "table_headers": ["Ingredient", "Qty Change", "Type", "Date", "Notes"],
        "table_rows": [[t.inventory_item.ingredient.name, t.quantity_change, t.get_transaction_type_display(), localtime(t.date).strftime('%Y-%m-%d %H:%M'), t.notes, t.id] for t in transactions],
        "inventory_items": InventoryItem.objects.select_related('ingredient').all(),
    }
    return render(request, "inventory/inventory_transactions.html", context)


def delete_inventory_transaction(request, pk):
    get_object_or_404(InventoryTransaction, pk=pk).delete()
    return redirect("inventory_trans")