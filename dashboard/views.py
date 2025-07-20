from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect,  get_object_or_404
from inventory.models import InventoryItem
from bookings.models import Booking, BookingStatus

from rooms.models import RoomType, Room
from accounts.models import User
import csv
import io
import datetime
import openpyxl
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from nepali_datetime import date as nep_date, datetime as nep_datetime
from io import BytesIO
from reportlab.pdfgen import canvas
from django.utils.text import slugify

from resorts.models import Resort
import pandas as pd
from reportlab.lib.pagesizes import letter
from django.contrib.auth.decorators import login_required

@login_required
def index(request):
    return render(request, 'index.html')

@login_required
def booking(request):
    bookings = Booking.objects.select_related('room', 'resort').all()

    if request.method == "POST":
        room_id = request.POST.get("room")
        resort_id = request.POST.get("resort")
        check_in = request.POST.get("check-in")
        check_out = request.POST.get("check-out")
        total_price = request.POST.get("total-price")
        status = request.POST.get("status")

        Booking.objects.create(
            user=request.user,
            resort_id=resort_id,
            room_id=room_id,
            check_in=check_in,
            check_out=check_out,
            total_price=total_price,
            status=status
        )
        return redirect("booking")

    # Export logic
    export = request.GET.get("export")
    if export == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="room_bookings.csv"'
        writer = csv.writer(response)
        writer.writerow(["Room", "Check In", "Check Out", "Total Price", "Status"])
        for b in bookings:
            writer.writerow([b.room.name, b.check_in, b.check_out, b.total_price, b.status])
        return response

    elif export == "excel":
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="room_bookings.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Bookings')

        columns = ["Room", "Check In", "Check Out", "Total Price", "Status"]
        for col_num in range(len(columns)):
            ws.write(0, col_num, columns[col_num])

        for row_num, b in enumerate(bookings, 1):
            ws.write(row_num, 0, b.room.name)
            ws.write(row_num, 1, str(b.check_in))
            ws.write(row_num, 2, str(b.check_out))
            ws.write(row_num, 3, float(b.total_price))
            ws.write(row_num, 4, b.status)

        wb.save(response)
        return response

    elif export == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="room_bookings.pdf"'

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)

        y = 750
        p.drawString(100, y, "Room Bookings")
        y -= 20

        for b in bookings:
            p.drawString(100, y, f"Room: {b.room.name} | {b.check_in} to {b.check_out} | NPR {b.total_price} | {b.status}")
            y -= 15

        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    context = {
        "page_title": "Room Booking",
        "table_headers": ["Room", "Check In", "Check Out", "Total Price", "Status"],
        "table_rows": [
            [b.room.name, b.check_in, b.check_out, b.total_price, b.get_status_display()] for b in bookings
        ],
        "form_fields": [
            {"label": "Room", "name": "room", "type": "select", "choices": Room.objects.all()},
            {"label": "Resort", "name": "resort", "type": "select", "choices": Resort.objects.all()},
            {"label": "Check In","class": "nepali-datepicker","name": "check-in", "type": "date"},
            {"label": "Check Out", "class": "nepali-datepicker","name": "check-out", "type": "nepali-date"},
            {"label": "Total Price", "name": "total-price", "type": "text"},
            {"label": "Status", "name": "status", "type": "select", "choices": BookingStatus.choices},
        ],
    }
    return render(request, "bookings/booking.html", context)

@login_required
def delete_booking(request, pk):
    booking = get_object_or_404(Booking, pk=pk)
    booking.delete()
    return redirect("booking")

def export_csv(bookings):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=bookings.csv"
    writer = csv.writer(response)
    writer.writerow(["Room", "Check In", "Check Out", "Total Price", "Status"])
    for b in bookings:
        writer.writerow([b.room.name, b.check_in, b.check_out, b.total_price, b.get_status_display()])
    return response

def export_excel(bookings):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Room", "Check In", "Check Out", "Total Price", "Status"])
    for b in bookings:
        ws.append([b.room.name, str(b.check_in), str(b.check_out), float(b.total_price), b.get_status_display()])
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="bookings.xlsx"'
    wb.save(response)
    return response

def export_pdf(bookings):
    # Placeholder - PDF generation (e.g., ReportLab or WeasyPrint)
    response = HttpResponse("PDF generation not implemented yet", content_type="text/plain")
    return response


def inventory_items(request):
    context = {
        "page_title": "Inventory Items",
        "table_headers": ["Item", "Stock", "Unit"],
        "table_rows": [
            ["Rice", "100", "kg"],
            ["Sugar", "50", "kg"],
        ]
    }
    return render(request,'bookings/booking.html', context)

def inventory_trans(request):
    context = {
        "page_title": "Inventory Transactions",
        "table_headers": ["Transaction ID", "Item", "Quantity", "Type", "Date"],
        "table_rows": [
            ["TXN001", "Rice", "20", "Out", "2025-07-01"],
            ["TXN002", "Oil", "15", "In", "2025-07-02"],
        ]
    }
    return render(request, "bookings/booking.html", context)

def kot_items(request):
    context = {
        "page_title": "KOT Items",
        "table_headers": ["Item", "Table No", "Quantity", "St6atus"],
        "table_rows": [
            ["Momo", "Table 5", "3", "Served"],
            ["Pizza", "Table 2", "1", "Preparing"],
        ]
    }
    return render(request, "bookings/booking.html", context)

def menu_items(request):
    context = {
        "page_title": "Menu Items",
        "table_headers": ["Name", "Category", "Price", "Availability"],
        "table_rows": [
            ["Burger", "Fast Food", "$5", "Available"],
            ["Pasta", "Main Course", "$7", "Out of Stock"],
        ]
    }
    return render(request, "bookings/booking.html", context)

def room_type(request):
    queryset = RoomType.objects.all()
    search = request.GET.get("search")

    if search:
        queryset = queryset.filter(name__icontains=search)

    export_type = request.GET.get("export")
    if export_type == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="room_types.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Description'])
        for obj in queryset:
            writer.writerow([obj.name, obj.description])
        return response

    elif export_type == "excel":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Description'])
        for obj in queryset:
            sheet.append([obj.name, obj.description])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="room_types.xlsx"'
        workbook.save(response)
        return response

    elif export_type == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="room_types.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.drawString(100, y, 'Room Types')
        y -= 20
        for obj in queryset:
            p.drawString(100, y, f"{obj.name} - {obj.description}")
            y -= 20
        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    if request.method == "POST":
        obj_id = request.POST.get("id")
        name = request.POST.get("type")
        description = request.POST.get("description")
        if obj_id:
            obj = RoomType.objects.get(id=obj_id)
            obj.name = name
            obj.description = description
            obj.save()
        else:
            RoomType.objects.create(name=name, description=description)
        return redirect("room_type")

    context = {
        "page_title": "Room Types",
        "table_headers": ["Type", "Description"],
        "table_rows": [[obj.name, obj.description, obj.id] for obj in queryset]
    }
    return render(request, "rooms/room_type.html", context)

def delete_room_type(request, pk):
    obj = get_object_or_404(RoomType, pk=pk)
    obj.delete()
    return redirect("room_type")



def room(request):
    queryset = Room.objects.select_related('resort', 'room_type').all()
    search = request.GET.get("search")

    if search:
        queryset = queryset.filter(name__icontains=search)

    export_type = request.GET.get("export")
    if export_type == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="rooms.csv"'
        writer = csv.writer(response)
        writer.writerow(['Resort', 'Room Type', 'Name', 'Description', 'Price/Night', 'Max Guests', 'Available'])
        for obj in queryset:
            writer.writerow([
                obj.resort.name,
                obj.room_type.name if obj.room_type else '',
                obj.name,
                obj.description,
                obj.price_per_night,
                obj.max_guests,
                "Yes" if obj.is_available else "No",
            ])
        return response

    elif export_type == "excel":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Resort', 'Room Type', 'Name', 'Description', 'Price/Night', 'Max Guests', 'Available'])
        for obj in queryset:
            sheet.append([
                obj.resort.name,
                obj.room_type.name if obj.room_type else '',
                obj.name,
                obj.description,
                obj.price_per_night,
                obj.max_guests,
                "Yes" if obj.is_available else "No",
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="rooms.xlsx"'
        workbook.save(response)
        return response

    elif export_type == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rooms.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.drawString(100, y, 'Rooms')
        y -= 20
        for obj in queryset:
            p.drawString(100, y, f"{obj.resort.name} - {obj.name} - {obj.room_type.name if obj.room_type else ''}")
            y -= 20
        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    if request.method == "POST":
        obj_id = request.POST.get("id")
        resort_id = request.POST.get("resort")
        room_type_id = request.POST.get("room_type")
        name = request.POST.get("name")
        description = request.POST.get("description")
        price = request.POST.get("price")
        guests = request.POST.get("max_guests")
        available = request.POST.get("status") == "true"  # from select field

        if obj_id:
            room = get_object_or_404(Room, pk=obj_id)
            room.resort_id = resort_id
            room.room_type_id = room_type_id
            room.name = name
            room.description = description
            room.price_per_night = price
            room.max_guests = guests
            room.is_available = available
            room.save()
        else:
            Room.objects.create(
                resort_id=resort_id,
                room_type_id=room_type_id,
                name=name,
                description=description,
                price_per_night=price,
                max_guests=guests,
                is_available=available
            )
        return redirect("room")

    context = {
        "page_title": "Rooms",
        "table_headers": ["Resort", "Room Type", "Name", "Price/Night", "Max Guests", "Available"],
        "table_rows": [[
            r.resort.name,
            r.room_type.name if r.room_type else '',
            r.name,
            r.price_per_night,
            r.max_guests,
            "Yes" if r.is_available else "No",
            r.id,
            r.description or '',
            r.resort.id,
            r.room_type.id if r.room_type else '',
            "true" if r.is_available else "false",
        ] for r in queryset],
        "resorts": Resort.objects.all(),
        "room_types": RoomType.objects.all(),
    }
    return render(request, "rooms/room.html", context)

def delete_room(request, pk):
    get_object_or_404(Room, pk=pk).delete()
    return redirect("room")