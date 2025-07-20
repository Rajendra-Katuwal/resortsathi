import csv
import openpyxl
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from reportlab.pdfgen import canvas
from .models import MenuItem, KOT, KOTItem
from django.utils.timezone import localtime
from bookings.models import Booking

def menu_items(request):
    items = MenuItem.objects.all()
    search = request.GET.get("search")

    if search:
        items = items.filter(name__icontains=search)

    # Export Section
    export = request.GET.get("export")
    if export == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="menu_items.csv"'
        writer = csv.writer(response)
        writer.writerow(["Name", "Description", "Price", "Available"])
        for item in items:
            writer.writerow([item.name, item.description, item.price, "Yes" if item.is_available else "No"])
        return response

    elif export == "excel":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Description", "Price", "Available"])
        for item in items:
            sheet.append([item.name, item.description, float(item.price), "Yes" if item.is_available else "No"])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="menu_items.xlsx"'
        workbook.save(response)
        return response

    elif export == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="menu_items.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.drawString(100, y, 'Menu Items')
        y -= 20
        for item in items:
            p.drawString(100, y, f"{item.name} - Rs {item.price} - {'Available' if item.is_available else 'Unavailable'}")
            y -= 15
        p.showPage()
        p.save()
        response.write(buffer.getvalue())
        buffer.close()
        return response

    # Add/Edit Form Handling
    if request.method == "POST":
        item_id = request.POST.get("id")
        name = request.POST.get("name")
        description = request.POST.get("description")
        price = request.POST.get("price")
        is_available = request.POST.get("is_available") == "True"

        if item_id:
            item = get_object_or_404(MenuItem, pk=item_id)
            item.name = name
            item.description = description
            item.price = price
            item.is_available = is_available
            item.save()
        else:
            MenuItem.objects.create(
                name=name,
                description=description,
                price=price,
                is_available=is_available,
            )
        return redirect("menu_items")

    context = {
        "page_title": "Menu Items",
        "table_headers": ["Name", "Description", "Price", "Available"],
        "table_rows": [[item.name, item.description, item.price, "Yes" if item.is_available else "No", item.id] for item in items],
    }
    return render(request, "kot/menu_items.html", context)

def delete_menu_item(request, pk):
    item = get_object_or_404(MenuItem, pk=pk)
    item.delete()
    return redirect("menu_items")

def kot_items(request):
    queryset = KOTItem.objects.select_related('kot', 'menu_item').all()

    search = request.GET.get("search")
    if search:
        queryset = queryset.filter(menu_item__name__icontains=search)

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="kot_items.csv"'
        writer = csv.writer(response)
        writer.writerow(["KOT", "Menu Item", "Quantity"])
        for obj in queryset:
            writer.writerow([obj.kot.id, obj.menu_item.name, obj.quantity])
        return response

    elif request.GET.get("export") == "excel":
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["KOT", "Menu Item", "Quantity"])
        for obj in queryset:
            sheet.append([obj.kot.id, obj.menu_item.name, obj.quantity])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="kot_items.xlsx"'
        wb.save(response)
        return response

    elif request.GET.get("export") == "pdf":
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        p.drawString(100, 800, "KOT Items")
        y = 780
        for obj in queryset:
            p.drawString(100, y, f"KOT #{obj.kot.id} - {obj.quantity} x {obj.menu_item.name}")
            y -= 20
        p.showPage()
        p.save()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="kot_items.pdf"'
        response.write(buffer.getvalue())
        buffer.close()
        return response

    if request.method == "POST":
        kot_id = request.POST.get("kot")
        menu_item_id = request.POST.get("menu_item")
        quantity = request.POST.get("quantity")
        obj_id = request.POST.get("id")

        if obj_id:
            obj = get_object_or_404(KOTItem, pk=obj_id)
            obj.kot_id = kot_id
            obj.menu_item_id = menu_item_id
            obj.quantity = quantity
            obj.save()
        else:
            KOTItem.objects.create(kot_id=kot_id, menu_item_id=menu_item_id, quantity=quantity)
        return redirect("kot_items")

    context = {
        "page_title": "KOT Items",
        "table_headers": ["KOT", "Menu Item", "Quantity"],
        "table_rows": [[obj.kot.id, obj.menu_item.name, obj.quantity, obj.id] for obj in queryset],
        "kots": KOT.objects.all(),
        "menu_items": MenuItem.objects.all(),
    }
    return render(request, "kot/kot_items.html", context)

def delete_kot_item(request, pk):
    get_object_or_404(KOTItem, pk=pk).delete()
    return redirect("kot_items")

def kot_list(request):
    query = request.GET.get("search")
    kot_qs = KOT.objects.select_related("booking")
    if query:
        kot_qs = kot_qs.filter(booking__id__icontains=query)

    if request.method == "POST":
        pk = request.POST.get("id")
        booking_id = request.POST.get("booking")
        status = request.POST.get("status")

        if pk:
            kot = get_object_or_404(KOT, pk=pk)
            kot.booking_id = booking_id
            kot.status = status
            kot.save()
        else:
            KOT.objects.create(booking_id=booking_id, status=status)

        return redirect("kot")

    export = request.GET.get("export")
    if export == "csv":
        response = HttpResponse(content_type="text/csv")
        response['Content-Disposition'] = 'attachment; filename="kots.csv"'
        writer = csv.writer(response)
        writer.writerow(["KOT ID", "Booking ID", "Status", "Created At"])
        for kot in kot_qs:
            writer.writerow([kot.id, kot.booking.id, kot.get_status_display(), localtime(kot.created_at).strftime("%Y-%m-%d %H:%M")])
        return response

    elif export == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["KOT ID", "Booking ID", "Status", "Created At"])
        for kot in kot_qs:
            ws.append([kot.id, kot.booking.id, kot.get_status_display(), localtime(kot.created_at).strftime("%Y-%m-%d %H:%M")])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="kots.xlsx"'
        wb.save(response)
        return response

    elif export == "pdf":
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = 'attachment; filename="kots.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        p.drawString(100, 800, "KOT Records")
        y = 780
        for kot in kot_qs:
            p.drawString(100, y, f"KOT #{kot.id} | Booking: {kot.booking.id} | Status: {kot.get_status_display()}")
            y -= 15
        p.showPage()
        p.save()
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response

    context = {
        "page_title": "KOT",
        "table_rows": [
            [kot.id, kot.booking.id, kot.get_status_display(), localtime(kot.created_at).strftime("%Y-%m-%d %H:%M")] for kot in kot_qs
        ],
        "bookings": Booking.objects.all(),
        "status_choices": KOT.status_choices,
    }
    return render(request, "kot/kot.html", context)

def delete_kot(request, pk):
    get_object_or_404(KOT, pk=pk).delete()
    return redirect("kot")

from django.shortcuts import render, redirect
from .models import KOT, KOTItem, MenuItem
from bookings.models import Booking

def create_kot(request):
    if request.method == 'POST':
        booking_type = request.POST.get('booking_type')
        booking_id = request.POST.get('booking_id') if booking_type == 'booking' else None
        status = request.POST.get('status')
        menu_item_ids = request.POST.getlist('menu_item_ids[]')
        quantities = request.POST.getlist('quantities[]')

        booking = Booking.objects.get(id=booking_id) if booking_id else None
        kot = KOT.objects.create(booking=booking, status=status)

        for item_id, qty in zip(menu_item_ids, quantities):
            if item_id and qty:
                KOTItem.objects.create(
                    kot=kot,
                    menu_item_id=item_id,
                    quantity=int(qty)
                )

        return redirect('kot_list')  # Redirect to list page

    bookings = Booking.objects.all()
    menu_items = MenuItem.objects.filter(is_available=True)
    status_choices = KOT.status_choices
    return render(request, 'kot/create_kot.html', {
        'bookings': bookings,
        'menu_items': menu_items,
        'status_choices': status_choices
    })

def kot(request):
    kots = KOT.objects.all().order_by('-created_at')
    return render(request, 'kot/kot_list.html', {'kots': kots})
