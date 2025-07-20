from django.shortcuts import render
from .models import *
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from datetime import date

def category_list(request):
    categories = Category.objects.all()
    return render(request, 'accounting/category_list.html', {'categories': categories})

def transaction_list(request):
    transactions = Transaction.objects.all()
    return render(request, 'accounting/transaction_list.html', {'transactions': transactions})

def invoice_list(request):
    invoices = Invoice.objects.all()
    return render(request, 'accounting/invoice_list.html', {'invoices': invoices})

def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    category.delete()
    messages.success(request, "Category deleted successfully.")
    return redirect('category_list')

def get_ledger(account_id, start_date, end_date):
    items = JournalItem.objects.filter(
        account_id=account_id,
        journal__date__range=[start_date, end_date]
    )
    return items.order_by('journal__date')

def trial_balance(start_date, end_date):
    accounts = Account.objects.all()
    data = []
    for account in accounts:
        items = JournalItem.objects.filter(
            account=account,
            journal__date__range=[start_date, end_date]
        )
        debit = sum([i.debit for i in items])
        credit = sum([i.credit for i in items])
        data.append({'account': account, 'debit': debit, 'credit': credit})
    return data

def income_statement(start_date, end_date):
    income_accounts = Account.objects.filter(type='income')
    expense_accounts = Account.objects.filter(type='expense')

    total_income = sum([
        sum(i.credit for i in JournalItem.objects.filter(account=a, journal__date__range=[start_date, end_date]))
        for a in income_accounts
    ])

    total_expense = sum([
        sum(i.debit for i in JournalItem.objects.filter(account=a, journal__date__range=[start_date, end_date]))
        for a in expense_accounts
    ])

    return {
        'total_income': total_income,
        'total_expense': total_expense,
        'net_profit': total_income - total_expense
    }

def balance_sheet(date):
    assets = Account.objects.filter(type='asset')
    liabilities = Account.objects.filter(type='liability')
    equity = Account.objects.filter(type='equity')

    def account_balance(account):
        debit = sum(i.debit for i in JournalItem.objects.filter(account=account, journal__date__lte=date))
        credit = sum(i.credit for i in JournalItem.objects.filter(account=account, journal__date__lte=date))
        return debit - credit

    return {
        'assets': [(a, account_balance(a)) for a in assets],
        'liabilities': [(l, account_balance(l)) for l in liabilities],
        'equity': [(e, account_balance(e)) for e in equity],
    }

def balance_sheet_view(request):
    resort = Resort.objects.get(user=request.user)
    assets = []
    asset_accounts = Account.objects.filter(type='asset')
    for acc in asset_accounts:
        debit = JournalItem.objects.filter(account=acc).aggregate(total=models.Sum('debit'))['total'] or 0
        credit = JournalItem.objects.filter(account=acc).aggregate(total=models.Sum('credit'))['total'] or 0
        assets.append((acc, debit - credit))

    # Liabilities
    liabilities = []
    liability_accounts = Account.objects.filter(type='liability')
    for acc in liability_accounts:
        debit = JournalItem.objects.filter(account=acc).aggregate(total=models.Sum('debit'))['total'] or 0
        credit = JournalItem.objects.filter(account=acc).aggregate(total=models.Sum('credit'))['total'] or 0
        liabilities.append((acc, credit - debit))

    # Equity
    equity = []
    equity_accounts = Account.objects.filter(type='equity')
    for acc in equity_accounts:
        debit = JournalItem.objects.filter(account=acc).aggregate(total=models.Sum('debit'))['total'] or 0
        credit = JournalItem.objects.filter(account=acc).aggregate(total=models.Sum('credit'))['total'] or 0
        equity.append((acc, credit - debit))

    return render(request, 'accounting/balance_sheet.html', {
        'resort': resort,
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity,
        'date': date.today(),
    })

# accounting/views.py

def income_statement_view(request):
    resort = Resort.objects.get(user=request.user)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        today = date.today()
        start_date = date(today.year, 1, 1)
        end_date = today

    income_accounts = Account.objects.filter(type='income')
    expense_accounts = Account.objects.filter(type='expense')

    income_data = []
    expense_data = []
    total_income = 0
    total_expense = 0

    for acc in income_accounts:
        credit = JournalItem.objects.filter(account=acc, journal__date__range=[start_date, end_date]).aggregate(total=models.Sum('credit'))['total'] or 0
        income_data.append((acc, credit))
        total_income += credit

    for acc in expense_accounts:
        debit = JournalItem.objects.filter(account=acc, journal__date__range=[start_date, end_date]).aggregate(total=models.Sum('debit'))['total'] or 0
        expense_data.append((acc, debit))
        total_expense += debit

    net_profit = total_income - total_expense

    return render(request, 'accounting/income_statement.html', {
        'resort': resort,
        'start_date': start_date,
        'end_date': end_date,
        'income_data': income_data,
        'expense_data': expense_data,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_profit': net_profit
    })
def ledger_view(request, account_id):
    resort = request.user.resort
    account = get_object_or_404(Account, id=account_id)
    items = JournalItem.objects.filter(account=account).order_by('journal__date')

    balance = 0
    ledger_data = []
    for item in items:
        if account.type in ['asset', 'expense']:
            balance += item.debit - item.credit
        else:
            balance += item.credit - item.debit
        ledger_data.append({
            'date': item.journal.date,
            'description': item.journal.description,
            'debit': item.debit,
            'credit': item.credit,
            'balance': balance,
        })

    return render(request, 'accounting/ledger.html', {
        'resort': resort,
        'account': account,
        'ledger_data': ledger_data
    })
import openpyxl
from django.http import HttpResponse
def export_balance_sheet_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    ws.append(["Particulars (Liabilities & Capital)", "Amount", "Particulars (Assets)", "Amount"])

    # Fill from your balance sheet logic here...

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=balance_sheet.xlsx'
    wb.save(response)
    return response

# accounting/views.py
from django.http import HttpResponse
from .utils.pdf import render_to_pdf

def balance_sheet_pdf(request):
    # Use same data structure as your HTML view
    context = generate_balance_sheet_context(request)
    pdf = render_to_pdf("accounting/balance_sheet_pdf.html", context)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="balance_sheet.pdf"'
    return response

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
import io
import xlsxwriter
from django.template.loader import get_template
from xhtml2pdf import pisa

from .models import Account, Transaction  # Update with your actual models

# ---- Income Statement PDF ----
def income_statement_pdf(request):
    template_path = 'accounting/reports/income_statement_pdf.html'
    context = {
        'company': "ResortSathi",
        'income': 10000,  # Replace with dynamic content
        'expenses': 5000,
        'net_income': 5000,
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="income_statement.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('PDF generation failed')
    return response

# ---- Income Statement Excel ----
def export_income_statement_excel(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    # Sample data
    worksheet.write('A1', 'Income')
    worksheet.write('B1', 10000)
    worksheet.write('A2', 'Expenses')
    worksheet.write('B2', 5000)
    worksheet.write('A3', 'Net Income')
    worksheet.write('B3', 5000)

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="income_statement.xlsx"'
    return response

# ---- Ledger PDF ----
def ledger_pdf(request, account_id):
    account = get_object_or_404(Account, id=account_id)
    transactions = Transaction.objects.filter(account=account)

    template_path = 'accounting/reports/ledger_pdf.html'
    context = {
        'account': account,
        'transactions': transactions,
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename="ledger_{account.name}.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('PDF generation failed')
    return response

# ---- Ledger Excel ----
def ledger_excel(request, account_id):
    account = get_object_or_404(Account, id=account_id)
    transactions = Transaction.objects.filter(account=account)

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    worksheet.write('A1', 'Date')
    worksheet.write('B1', 'Description')
    worksheet.write('C1', 'Amount')

    row = 1
    for t in transactions:
        worksheet.write(row, 0, str(t.date))
        worksheet.write(row, 1, t.description)
        worksheet.write(row, 2, float(t.amount))
        row += 1

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ledger_{account.name}.xlsx"'
    return response
# views.py
def ledger_list(request):
    # your logic here
    return render(request, 'accounting/ledger_list.html')
